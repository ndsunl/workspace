import openpyxl


def main():
    """
    单元格内换行
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.title = 'WorkWheetTitle'

    ws['A1'] = 'A\nB\nC'
    ws['A1'].alignment = openpyxl.styles.Alignment(wrapText=True)

    wb.save('example.xlsx')


if __name__ == '__main__':
    main()