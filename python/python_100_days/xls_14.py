import openpyxl


def main():
    """
    插入超链接
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.title = 'WorkSheelTitle'
    ws2 = wb.create_sheet('Example')

    # 设置超链接，并将鼠标定格在A5单元格
    ws['A1'] = "Link"
    ws['A1'].hyperlink = 'example.xlsx#Example!A5'

    wb.save('example.xlsx')


if __name__ == '__main__':
    main()