import openpyxl
from openpyxl.styles import PatternFill


def main():
    """
    单元格填充颜色
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.title = 'WorkSheetTitle'

    fill = PatternFill(
        fill_type='solid',
        fgColor='FFFF0000'
    )
    b2 = ws['B2']
    b2.fill = fill
    b2.value = 'TEST'

    wb.save('example.xlsx')


if __name__ == '__main__':
    main()