import openpyxl


def main():
    """
    合并单元格
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.title = "WorkSheetTitle"

    ws.merge_cells("A1:E1")
    ws['A1'] = 'HOGE'

    wb.save('example.xlsx')


if __name__ == "__main__":
    main()