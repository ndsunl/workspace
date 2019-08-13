import openpyxl

def main():
    wb = openpyxl.Workbook()
    # 当毅页
    ws = wb.active

    # 更改默认名称
    ws.title = 'WorkSheetTitle'

    # 定义第二个页
    ws2 = wb.create_sheet("NewWorkSheet2")

    # 定义第三个页
    # 0 的设定让该页置于最前
    ws3 = wb.create_sheet('NewWorkSheet3', 0)

    # 保存
    wb.save('example.xlsx')


if __name__ == '__main__':
    main()