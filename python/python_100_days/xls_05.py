import openpyxl


def main():
    """指定坐标赋值"""
    wb = openpyxl.Workbook()
    ws = wb.active

    # 更改表格名称
    ws.title = "WorkSheetTitle"

    # 给单元格赋值
    ws['A1'] = 'HOGE'
    ws['B1'] = 'FUGA'

    # 保存
    wb.save('example.xlsx')


if __name__ == '__main__':
    main()